"""Excel parsing primitives shared by all sheet parsers.

Responsibilities:
    - Open workbook with rich-text awareness.
    - Walk every cell of a sheet yielding `CellInfo` records with
      normalized color, font, indent and merge-anchor information.
    - Resolve color values across openpyxl's three representations
      (rgb / indexed / theme) into a single `#RRGGBB` string.
    - Infer per-row indent level based on the left-most non-empty
      column observed across the sheet (this is how the sample
      仕様書 encodes hierarchy).
    - Build a plain-text dump and a markdown rendering of the sheet
      for fallback / debug consumers.

Nothing here knows anything about the *meaning* of a sheet — that
is the responsibility of `parsers/*.py`.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

from loguru import logger
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Default Excel indexed palette (OOXML spec) — only used when a Color object
# is of type 'indexed'. Source: https://openpyxl.readthedocs.io
# ---------------------------------------------------------------------------
_INDEXED_PALETTE: list[str] = [
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
    "800000", "008000", "000080", "808000", "800080", "008080", "C0C0C0", "808080",
    "9999FF", "993366", "FFFFCC", "CCFFFF", "660066", "FF8080", "0066CC", "CCCCFF",
    "000080", "FF00FF", "FFFF00", "00FFFF", "800080", "800000", "008080", "0000FF",
    "00CCFF", "CCFFFF", "CCFFCC", "FFFF99", "99CCFF", "FF99CC", "CC99FF", "FFCC99",
    "3366FF", "33CCCC", "99CC00", "FFCC00", "FF9900", "FF6600", "666699", "969696",
    "003366", "339966", "003300", "333300", "993300", "993366", "333399", "333333",
    "000000", "FFFFFF",
]


@dataclass
class TextRun:
    """A single styled run inside a cell (used for rich text)."""
    text: str
    color: str | None = None          # '#RRGGBB' or None
    bold: bool = False
    italic: bool = False
    underline: bool = False


@dataclass
class CellInfo:
    row: int
    col: int                          # 1-based column number
    col_letter: str
    text: str                         # full cell text (concatenated runs)
    runs: list[TextRun] = field(default_factory=list)
    font_color: str | None = None     # dominant / first run color
    fill_color: str | None = None     # cell background ('#RRGGBB' or None)
    bold: bool = False
    italic: bool = False
    font_size: float | None = None
    is_merge_anchor: bool = False     # this is the top-left of a merged range
    merged_range: str | None = None   # e.g. 'B5:K10' if anchor; else None
    hyperlink: str | None = None
    indent_level: int = 0             # filled in by `infer_indent_levels`


@dataclass
class SheetSnapshot:
    """The output of `read_sheet`: every cell + structural metadata."""
    name: str
    index: int
    cells: list[CellInfo]
    merged_ranges: list[str]
    max_row: int
    max_col: int
    has_images: bool


# ---------------------------------------------------------------------------
# Color resolution
# ---------------------------------------------------------------------------

def _resolve_color(c: Color | None, theme_lookup: dict[int, str] | None = None) -> str | None:
    """Best-effort convert an openpyxl Color to '#RRGGBB'.

    Returns None if the color is the default / unset / 'auto'.
    Theme colors are returned as None unless a `theme_lookup` mapping is
    provided. We don't compute the theme dynamically here because doing
    so reliably requires parsing the workbook's themeN.xml. For our
    purposes (detecting JIRA-tag colors in 仕様書), the colors of
    interest are always set as explicit RGB.
    """
    if c is None:
        return None
    t = getattr(c, "type", None)
    if t == "rgb":
        rgb = c.rgb
        if not rgb or rgb in ("00000000",):
            return None
        # rgb is ARGB; drop the alpha
        if len(rgb) == 8:
            rgb = rgb[2:]
        return f"#{rgb.upper()}"
    if t == "indexed":
        idx = c.indexed
        if idx is None or idx >= len(_INDEXED_PALETTE):
            return None
        return f"#{_INDEXED_PALETTE[idx].upper()}"
    if t == "theme":
        if theme_lookup is not None and c.theme in theme_lookup:
            return theme_lookup[c.theme]
        return None
    return None


def _cell_runs(cell: Cell) -> tuple[str, list[TextRun]]:
    """Extract text + per-run formatting from a cell."""
    val = cell.value
    if val is None:
        return "", []

    # Rich text — multiple runs in one cell
    if isinstance(val, CellRichText):
        runs: list[TextRun] = []
        parts: list[str] = []
        for piece in val:
            if isinstance(piece, str):
                runs.append(TextRun(text=piece))
                parts.append(piece)
            elif isinstance(piece, TextBlock):
                color = _resolve_color(piece.font.color) if piece.font else None
                bold = bool(piece.font and piece.font.b)
                italic = bool(piece.font and piece.font.i)
                underline = bool(piece.font and piece.font.u)
                runs.append(TextRun(
                    text=piece.text,
                    color=color,
                    bold=bold,
                    italic=italic,
                    underline=underline,
                ))
                parts.append(piece.text)
        return "".join(parts), runs

    # Plain value
    text = str(val).strip("　 \t\r\n")  # also strip JP full-width space
    if not text:
        return "", []
    color = _resolve_color(cell.font.color) if cell.font else None
    run = TextRun(
        text=text,
        color=color,
        bold=bool(cell.font and cell.font.b),
        italic=bool(cell.font and cell.font.i),
        underline=bool(cell.font and cell.font.u),
    )
    return text, [run]


def _dominant_color(runs: list[TextRun]) -> str | None:
    """Pick a representative color for the cell.

    Strategy: first run that has a non-default color. If all runs are
    plain, return None. This is good enough for tagging — for "any run
    has color X", iterate `runs` directly.
    """
    for r in runs:
        if r.color:
            return r.color
    return None


# ---------------------------------------------------------------------------
# Workbook / sheet readers
# ---------------------------------------------------------------------------

def open_workbook(path: str | Path) -> Workbook:
    """Open .xlsx with rich-text retention enabled."""
    return load_workbook(filename=str(path), data_only=False, rich_text=True)


def read_sheet(ws: Worksheet, index: int) -> SheetSnapshot:
    """Walk a sheet and return a SheetSnapshot.

    Empty cells (no value) are skipped to keep the snapshot compact.
    `infer_indent_levels` is called once at the end so each CellInfo
    knows its row's indent depth.
    """
    merged_ranges = [str(r) for r in ws.merged_cells.ranges]
    merged_anchors: dict[tuple[int, int], str] = {}
    for mr in ws.merged_cells.ranges:
        merged_anchors[(mr.min_row, mr.min_col)] = str(mr)

    cells: list[CellInfo] = []
    for row in ws.iter_rows():
        for cell in row:
            text, runs = _cell_runs(cell)
            if not text and cell.fill is None:
                continue
            if not text:
                # Skip purely-empty cells even if they have a fill —
                # they rarely carry content, just visual padding.
                continue
            fill_color = None
            if cell.fill and cell.fill.fgColor is not None:
                fill_color = _resolve_color(cell.fill.fgColor)
                # Excel's default 'no fill' often shows up as white with patternType=None
                if cell.fill.patternType is None:
                    fill_color = None

            anchor = merged_anchors.get((cell.row, cell.column))
            info = CellInfo(
                row=cell.row,
                col=cell.column,
                col_letter=get_column_letter(cell.column),
                text=text,
                runs=runs,
                font_color=_dominant_color(runs),
                fill_color=fill_color,
                bold=any(r.bold for r in runs) or bool(cell.font and cell.font.b),
                italic=any(r.italic for r in runs) or bool(cell.font and cell.font.i),
                font_size=float(cell.font.size) if cell.font and cell.font.size else None,
                is_merge_anchor=anchor is not None,
                merged_range=anchor,
                hyperlink=str(cell.hyperlink.target) if cell.hyperlink else None,
            )
            cells.append(info)

    infer_indent_levels(cells)
    return SheetSnapshot(
        name=ws.title,
        index=index,
        cells=cells,
        merged_ranges=merged_ranges,
        max_row=ws.max_row or 0,
        max_col=ws.max_column or 0,
        has_images=bool(getattr(ws, "_images", None)),
    )


def iter_sheets(wb: Workbook) -> Iterator[SheetSnapshot]:
    for idx, ws in enumerate(wb.worksheets):
        yield read_sheet(ws, idx)


# ---------------------------------------------------------------------------
# Hierarchy inference
# ---------------------------------------------------------------------------

def infer_indent_levels(cells: list[CellInfo]) -> None:
    """Assign `indent_level` to each cell, mutating in place.

    Heuristic: for each row, find the left-most occupied column (its
    `col`). The set of distinct left-most columns across the sheet,
    sorted ascending, becomes the "indent ladder" — row N's indent
    level is the ladder index of its row's left-most column.

    This handles the common 仕様書 pattern of indenting by *empty
    columns* rather than by space characters, where rows might start
    at A, B, C, D, ... corresponding to nesting depth 0, 1, 2, 3.
    """
    if not cells:
        return

    leftmost_by_row: dict[int, int] = {}
    for c in cells:
        prev = leftmost_by_row.get(c.row)
        if prev is None or c.col < prev:
            leftmost_by_row[c.row] = c.col

    ladder = sorted(set(leftmost_by_row.values()))
    rank = {col: i for i, col in enumerate(ladder)}

    for c in cells:
        c.indent_level = rank[leftmost_by_row[c.row]]


# ---------------------------------------------------------------------------
# Dump helpers (debugging + fallback parser)
# ---------------------------------------------------------------------------

def dump_text(snap: SheetSnapshot) -> str:
    """Plain-text dump of a sheet, one row per line.

    Cells on the same row are joined with TABs; rows are joined with
    newlines. Useful for FTS fallback and for feeding LLM classifiers.
    """
    rows: dict[int, list[CellInfo]] = {}
    for c in snap.cells:
        rows.setdefault(c.row, []).append(c)

    lines: list[str] = []
    for row_idx in sorted(rows):
        ordered = sorted(rows[row_idx], key=lambda x: x.col)
        lines.append("\t".join(c.text for c in ordered))
    return "\n".join(lines)


def dump_markdown(snap: SheetSnapshot, max_rows: int = 200) -> str:
    """A roughly-formatted markdown dump suitable for LLM prompts.

    Lines are prefixed with indent according to `indent_level`. This
    is *not* a faithful re-render of the spreadsheet — it just gives
    the LLM enough structure to reason about hierarchy.
    """
    rows: dict[int, list[CellInfo]] = {}
    for c in snap.cells:
        rows.setdefault(c.row, []).append(c)

    out: list[str] = [f"# Sheet: {snap.name}"]
    for i, row_idx in enumerate(sorted(rows)):
        if i >= max_rows:
            out.append(f"... ({len(rows) - max_rows} more rows omitted)")
            break
        ordered = sorted(rows[row_idx], key=lambda x: x.col)
        indent = "  " * ordered[0].indent_level
        text = " | ".join(c.text for c in ordered)
        out.append(f"{indent}- {text}")
    return "\n".join(out)


# ---------------------------------------------------------------------------
# Convenience
# ---------------------------------------------------------------------------

def load_and_iter(path: str | Path) -> Iterator[SheetSnapshot]:
    """One-shot: open file and yield every sheet snapshot. Caller
    is responsible for closing nothing — openpyxl handles it."""
    wb = open_workbook(path)
    try:
        yield from iter_sheets(wb)
    finally:
        wb.close()
        logger.debug(f"Closed workbook {path}")
