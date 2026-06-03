"""Screen-spec (画面仕様) parser.

These sheets contain one or more UI screenshots (typically of the
target system's screen), often annotated with numbered red boxes
(①②③), plus surrounding cells of explanatory text. We:

    1. Extract every embedded image (via image_extract.extract_sheet_images).
    2. Ask the VL model to describe each image in detail: layout,
       controls, labels, red-box numbered annotations.
    3. Pull *all* textual cells from the sheet into a "context block"
       and group them with the image description.
    4. Emit one chunk per image, plus optionally one extra chunk for
       any side-panel cell text (left navigation menu, etc).
"""

from __future__ import annotations

import json
import re

from loguru import logger

from ingestion.excel_utils import SheetSnapshot
from ingestion.image_extract import extract_sheet_images
from ingestion.parsers.base import (
    BaseSheetParser,
    ExtractedImageInfo,
    ParsedChunk,
    ParserContext,
    ParserOutput,
)
from llm.ollama_client import ollama
from openpyxl import load_workbook


_VL_PROMPT = """这是一张系统 UI 画面截图（来自一份日文设计文档）。请仔细分析并输出 JSON：

{
  "screen_title": "画面标题（如有可见的标题文字）",
  "layout": "整体布局描述（如：左侧菜单 + 右侧主操作区）",
  "regions": [
    {
      "name": "区块名/功能用途",
      "controls": ["按钮/输入框/下拉框等控件，含可见文字"]
    }
  ],
  "annotations": [
    {
      "marker": "①/②/③ 或红框编号",
      "location": "位于哪个区块/控件附近",
      "description": "该标注指向的对象"
    }
  ],
  "notes": "其他值得注意的视觉细节"
}

只输出 JSON，不要附加任何文字。如果某项无法识别，用 null 或空数组。"""


def _all_cell_text(snap: SheetSnapshot) -> str:
    """Dump every cell on this sheet, ordered by row/col, joined with newlines.
    Used to give the chunk a searchable text body alongside the VL description."""
    rows: dict[int, list] = {}
    for c in snap.cells:
        rows.setdefault(c.row, []).append(c)
    out: list[str] = []
    for row_idx in sorted(rows):
        cells = sorted(rows[row_idx], key=lambda x: x.col)
        out.append(" ".join(c.text for c in cells))
    return "\n".join(out)


async def _describe_image(image_path) -> dict:
    try:
        raw = await ollama.vision(_VL_PROMPT, [image_path])
        txt = raw.strip()
        if txt.startswith("```"):
            txt = re.sub(r"^```[a-zA-Z]*", "", txt).rstrip("`").strip()
        return json.loads(txt)
    except Exception as e:
        logger.warning(f"VL describe failed for {image_path}: {e}")
        return {}


def _as_str(x) -> str:
    """Tolerant string coercion for VL outputs that sometimes nest dicts
    inside what should be string fields."""
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    if isinstance(x, (int, float, bool)):
        return str(x)
    if isinstance(x, dict):
        # Common: {"text": "..."} or {"value": "..."}
        for k in ("text", "value", "name", "description", "title"):
            if k in x:
                return _as_str(x[k])
        return json.dumps(x, ensure_ascii=False)
    if isinstance(x, list):
        return ", ".join(_as_str(e) for e in x)
    return str(x)


def _render_description_md(desc) -> str:
    """Defensive renderer: VL output may have wrong types per field.
    Any item that isn't a dict where we expect one is rendered as a
    plain string."""
    if not isinstance(desc, dict) or not desc:
        return "_(画面描述生成失败)_"
    parts: list[str] = []
    if t := _as_str(desc.get("screen_title")):
        parts.append(f"**画面标题**: {t}")
    if l := _as_str(desc.get("layout")):
        parts.append(f"**布局**: {l}")

    regions = desc.get("regions") or []
    if isinstance(regions, list) and regions:
        parts.append("**区块与控件**:")
        for r in regions:
            if isinstance(r, dict):
                n = _as_str(r.get("name")) or "(unnamed)"
                controls = r.get("controls") or []
                parts.append(f"- {n}")
                if isinstance(controls, list):
                    for c in controls:
                        parts.append(f"  - {_as_str(c)}")
                elif controls:
                    parts.append(f"  - {_as_str(controls)}")
            else:
                parts.append(f"- {_as_str(r)}")

    anns = desc.get("annotations") or []
    if isinstance(anns, list) and anns:
        parts.append("**标注**:")
        for a in anns:
            if isinstance(a, dict):
                marker = _as_str(a.get("marker")) or "?"
                loc = _as_str(a.get("location")) or "?"
                d = _as_str(a.get("description"))
                parts.append(f"- {marker} @ {loc} — {d}")
            else:
                parts.append(f"- {_as_str(a)}")

    if n := _as_str(desc.get("notes")):
        parts.append(f"**备注**: {n}")
    return "\n".join(parts)


class ScreenParser(BaseSheetParser):
    name = "screen"

    async def parse(self, ctx: ParserContext) -> ParserOutput:
        snap = ctx.snapshot
        # Re-open the workbook to access ws._images on the actual sheet
        # (the snapshot doesn't carry openpyxl image objects).
        wb = load_workbook(str(ctx.xlsx_path))
        if snap.name not in wb.sheetnames:
            wb.close()
            return ParserOutput(notes=[f"sheet '{snap.name}' missing on re-open"])

        ws = wb[snap.name]
        extracted = extract_sheet_images(ws, ctx.image_dir)
        wb.close()

        side_text = _all_cell_text(snap)
        images_out: list[ExtractedImageInfo] = []
        chunks: list[ParsedChunk] = []
        order = 0
        doc_jiras = ctx.document_metadata.get("jira_codes", [])

        if not extracted:
            # No images; produce a single text chunk from cells so the
            # sheet remains retrievable.
            if side_text.strip():
                chunks.append(ParsedChunk(
                    text=side_text,
                    markdown=f"### {snap.name}\n\n{side_text}",
                    metadata={
                        "kind": "screen_text_only",
                        "sheet": snap.name,
                        "jira_tags": doc_jiras,
                    },
                    hierarchy_path=snap.name,
                    order=order,
                ))
            return ParserOutput(chunks=chunks)

        for img in extracted:
            desc = await _describe_image(img.file_path)
            desc_md = _render_description_md(desc)
            images_out.append(ExtractedImageInfo(
                file_path=img.file_path,
                anchor_cell=img.anchor_cell,
                vl_description=desc_md,
                annotations={"raw": desc},
            ))

            # Searchable text combines the rendered description with the
            # surrounding cell text — that's what the user is most likely
            # to ask about. We tolerate any malformed VL output here.
            combined_text_parts: list[str] = []
            if isinstance(desc, dict):
                combined_text_parts.append(_as_str(desc.get("screen_title")))
                combined_text_parts.append(_as_str(desc.get("layout")))
                for r in desc.get("regions") or []:
                    if isinstance(r, dict):
                        combined_text_parts.append(_as_str(r.get("name")))
                        ctrls = r.get("controls") or []
                        if isinstance(ctrls, list):
                            for c in ctrls:
                                combined_text_parts.append(_as_str(c))
                        else:
                            combined_text_parts.append(_as_str(ctrls))
                    else:
                        combined_text_parts.append(_as_str(r))
                for a in desc.get("annotations") or []:
                    if isinstance(a, dict):
                        combined_text_parts.append(_as_str(a.get("description")))
                    else:
                        combined_text_parts.append(_as_str(a))
            combined_text_parts.append(side_text)
            text = " ".join(p for p in combined_text_parts if p)

            md = f"### {snap.name}\n\n{desc_md}\n\n---\n\n**画面周围文字**:\n\n{side_text}"
            chunks.append(ParsedChunk(
                text=text,
                markdown=md,
                metadata={
                    "kind": "screen",
                    "sheet": snap.name,
                    "image_anchor": img.anchor_cell,
                    "annotations": desc.get("annotations") or [],
                    "jira_tags": doc_jiras,
                },
                hierarchy_path=snap.name,
                order=order,
            ))
            order += 1

        return ParserOutput(chunks=chunks, images=images_out)
