"""Render each sheet of an Excel file to a PNG screenshot.

Pipeline:  xlsx -> (LibreOffice headless) -> pdf -> (pdf2image) -> png per page.

Why this approach?
    - 100% cross-platform (Linux dev, Win11 target).
    - No need for Excel COM automation.
    - One PDF page roughly maps to one print-page of one sheet.

Caveats:
    - Wide sheets paginate. We rely on the workbook's print settings; for
      best results, ingestion can pre-set `fitToPage=True` per sheet before
      calling this module. That's done in `prepare_for_screenshot`.
    - LibreOffice must be installed on the host (apt: libreoffice;
      poppler-utils for pdf2image).
"""

from __future__ import annotations

import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path

from loguru import logger
from openpyxl import load_workbook
from pdf2image import convert_from_path

from config import settings


@dataclass
class SheetScreenshot:
    sheet_name: str
    sheet_index: int
    page_index: int          # 0-based; >0 means this is sheet was paginated
    file_path: Path


def prepare_for_screenshot(xlsx_path: Path, out_path: Path) -> Path:
    """Open workbook, set `fitToPage` on every sheet, save copy to out_path.

    Returns out_path. This mutates a *copy*, never the original.
    """
    wb = load_workbook(str(xlsx_path))
    for ws in wb.worksheets:
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 0 = unlimited vertically
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    wb.save(str(out_path))
    return out_path


def _run_libreoffice_to_pdf(xlsx_path: Path, out_dir: Path) -> Path:
    """Invoke `soffice --headless --convert-to pdf` and return the PDF path."""
    bin_ = settings.libreoffice_bin
    if not shutil.which(bin_):
        raise RuntimeError(
            f"LibreOffice binary '{bin_}' not found. "
            "Install via `apt install libreoffice` or set LIBREOFFICE_BIN."
        )
    cmd = [
        bin_,
        "--headless",
        "--convert-to", "pdf",
        "--outdir", str(out_dir),
        str(xlsx_path),
    ]
    logger.debug(f"Running: {' '.join(cmd)}")
    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
    if proc.returncode != 0:
        raise RuntimeError(
            f"LibreOffice conversion failed: {proc.stderr.strip() or proc.stdout.strip()}"
        )
    pdf = out_dir / f"{xlsx_path.stem}.pdf"
    if not pdf.exists():
        raise RuntimeError(f"Expected PDF not produced at {pdf}")
    return pdf


def render_workbook_screenshots(
    xlsx_path: Path,
    out_dir: Path,
    dpi: int = 110,
) -> list[SheetScreenshot]:
    """Produce one PNG per page of the workbook PDF.

    The mapping between PDF page → sheet is *order-based*: LibreOffice
    emits sheets in workbook order, and each sheet contributes 1+ pages
    depending on its print settings. With `fitToWidth=1` (set by
    `prepare_for_screenshot`), most sheets fit on one page wide; vertical
    pagination still happens for very long sheets.

    Because we cannot perfectly tell where one sheet ends and the next
    begins from the PDF alone, this function returns *all* pages and
    leaves it to the caller to associate pages with sheets. The
    convenience method `pick_first_page_per_sheet` handles the common
    case where each sheet fits on a single page.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as td:
        td_path = Path(td)
        prepared = prepare_for_screenshot(xlsx_path, td_path / xlsx_path.name)
        pdf = _run_libreoffice_to_pdf(prepared, td_path)
        images = convert_from_path(str(pdf), dpi=dpi)

        results: list[SheetScreenshot] = []
        # Without page-to-sheet info we mark page_index globally; the
        # caller can refine via heuristics or pre-counting print pages.
        for i, img in enumerate(images):
            path = out_dir / f"{xlsx_path.stem}__page{i:03d}.png"
            img.save(path, "PNG")
            results.append(SheetScreenshot(
                sheet_name="",
                sheet_index=-1,
                page_index=i,
                file_path=path,
            ))
        logger.info(f"Rendered {len(results)} page(s) from {xlsx_path.name}")
        return results


def render_single_sheet(
    xlsx_path: Path,
    sheet_name: str,
    out_dir: Path,
    dpi: int = 110,
) -> SheetScreenshot | None:
    """Render only `sheet_name` by writing a single-sheet xlsx copy.

    Cheaper than rendering the whole workbook when you just need one
    screenshot (e.g. on-demand for the VL classifier).
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(str(xlsx_path))
    if sheet_name not in wb.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found in {xlsx_path}")
        return None

    # Remove all other sheets so the PDF has exactly one page (or a few
    # vertical pages for very long sheets).
    for name in list(wb.sheetnames):
        if name != sheet_name:
            del wb[name]

    # Apply fit-to-width
    ws = wb[sheet_name]
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    with tempfile.TemporaryDirectory() as td:
        td_path = Path(td)
        tmp_xlsx = td_path / f"_one_{xlsx_path.stem}.xlsx"
        wb.save(str(tmp_xlsx))
        pdf = _run_libreoffice_to_pdf(tmp_xlsx, td_path)
        images = convert_from_path(str(pdf), dpi=dpi)
        if not images:
            return None
        # Take the first page; if the sheet is long, downstream code can
        # request additional pages explicitly.
        path = out_dir / f"{xlsx_path.stem}__{_safe_name(sheet_name)}.png"
        images[0].save(path, "PNG")
        return SheetScreenshot(
            sheet_name=sheet_name,
            sheet_index=-1,
            page_index=0,
            file_path=path,
        )


def _safe_name(s: str) -> str:
    return "".join(c if c.isalnum() or c in "-_." else "_" for c in s)[:60]
